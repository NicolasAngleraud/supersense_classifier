import pickle
import random
from collections import defaultdict
import openpyxl
from openpyxl.styles import PatternFill


def normalization_id(full_id, lang):
    if full_id.startswith(lang + ":"):
        return full_id.removeprefix(lang + ":").strip('_')
    elif full_id.startswith("<http://kaiko.getalp.org/dbnary/"):
        return full_id.removeprefix("<http://kaiko.getalp.org/dbnary/" + lang + "/").removesuffix(">").strip('_')
    else:
        return full_id


def extract_labels(text):
    if text:
        labels = set()
        label = ""
        definition = text
        labels_to_be_extracted = False
        if text.strip().startswith("("):
            labels_to_be_extracted = True
        while labels_to_be_extracted:
            end = definition.find(')')
            label = definition[:end+1]
            labels.add(label[1:-1].lower())
            definition = definition.removeprefix(label).strip()
            if not definition.strip().startswith('('):
                labels_to_be_extracted = False
        return labels, definition
    else:
        return None, None


def extract_quote(left_delimiter='"', right_delimiter='"', text=""):
    quote = text
    stext = text.split("@fr")
    for el in stext:
        if "rdf:value" in el:
            el = el.replace("'", "’")
            # el = el.strip('\\')
            start = el.find('"')
            end = el.rfind('"')
            quote = el[start+1:end].strip()
            # quote = el
            break
    # quote = quote.removeprefix('[ rdf:value')
    # quote = quote.strip('" ')
    # if len(quote) > 2:
    #    quote = quote[1:-1]
    return quote


class Page:
    """
    A class used to represent a page of the wiktionary

    ...

    Attributes
    ----------
    id : str
        the id of the page
    entry_ids : list
        the list of ids of the lexical entries of the page

    Methods
    -------
    is_multi_entries_page()
        Returns True if the page has more than one entry, False otherwise
    """

    # constructor
    def __init__(self, id, entry_ids=[]):

        # instance variable
        # id of the page
        self.id = id

        self.entry_ids = entry_ids

    def is_multi_entries_page(self):
        return len(self.entry_ids) > 1

    def get_entry_ids(self):
        return self.entry_ids

    def __str__(self):
        return f"PAGE_{self.id}: {str(self.entry_ids)}"


class lexicalEntry:
    """
    A class used to represent a lexical entry of the wiktionary

    ...

    Attributes
    ----------
    id : str
        a formatted string to print out what the animal says
    sense_ids : str
        a formatted string to print out what the animal says
    cf_ids : str
        a formatted string to print out what the animal says
    pos : str
        a formatted string to print out what the animal says
    lemma : str
        a formatted string to print out what the animal says
    synonyms : str
        the name of the animal
    morpho : str
        the sound that the animal makes
    is_mwe : int
        the number of legs the animal has (default 4)

    Methods
    -------
    says(sound=None)
        Prints the animals name and what sound it makes
    """
    # constructor
    def __init__(self, id, pos="", lemma="", synonyms=set(), sense_ids=[], morpho=[], cf_ids=[], is_mwe=False):
        self.id = id
        self.sense_ids = sense_ids
        self.cf_ids = cf_ids
        self.pos = pos
        self.lemma = lemma
        self.synonyms = synonyms
        self.morpho = morpho
        self.is_mwe = is_mwe

    def is_multi_senses_entry(self):
        return len(self.sense_ids) > 1

    def is_of_grammatical_category(self, grammatical_cats):
        return self.pos[0] in grammatical_cats

    def __str__(self):
        return f"   ENTRY_{self.id}: {self.sense_ids}\n"

    def add_morpho(self, morpho):
        self.morpho = morpho

    def set_sense_ids(self, sense_ids):
        self.sense_ids = sense_ids

    def get_lemma(self):
        return self.lemma

    def remove_sense_id(self, sense_id):
        self.sense_ids.remove(sense_id)

    def get_sense_ids(self):
        return self.sense_ids

    def get_cf_ids(self):
        return self.cf_ids


class lexicalSense:
    """
    A class used to represent an Animal

    ...

    Attributes
    ----------
    says_str : str
        a formatted string to print out what the animal says
    name : str
        the name of the animal
    sound : str
        the sound that the animal makes
    num_legs : int
        the number of legs the animal has (default 4)

    Methods
    -------
    says(sound=None)
        Prints the animals name and what sound it makes
    """

    # constructor
    def __init__(self, id, definition="", examples=[()], synonyms=set(), labels=set()):
        # instance variable
        # id of the lexical sense
        self.id = id

        # instance variable
        # definition is a string which contains the definition of the lexical sense
        self.definition = definition

        # instance variable
        # examples is a list of pairs containing each example as a tokenized list of strings
        # and the index of position in the sentence of the word associated with the lexical sense as an integer value
        self.examples = examples

        self.synonyms = synonyms

        self.labels = labels

    def __str__(self):
        string = f"       SENSE_{self.id}: \nDEF:\n{self.definition}"
        if len(self.examples) == 0:
            string += "\n"
            return string
        else:
            string += "\nEXAMPLES:\n"
            for example in self.examples:
                string += str(example) + "\n"
        return string

    def add_example(self, example):
        self.examples.append(example)

    def add_definition(self, definition):
        self.definition = definition

    def add_synonym(self, synonym):
        self.synonyms.add(synonym)

    def add_label(self, label):
        self.labels.add(label)

    def get_definition(self):
        return self.definition

    def get_examples(self):
        return self.examples


class Wiktionary:
    """
    A class used to represent an Animal

    ...

    Attributes
    ----------
    says_str : str
        a formatted string to print out what the animal says
    name : str
        the name of the animal
    sound : str
        the sound that the animal makes
    num_legs : int
        the number of legs the animal has (default 4)

    Methods
    -------
    says(sound=None)
        Prints the animals name and what sound it makes
    """
    # constructor
    def __init__(self):
        self.pages = {}
        self.lexical_entries = {}
        self.lexical_senses = {}
        self.canonical_forms = {}

    def is_empty(self):
        return len(self.pages) == 0

    def get_pages(self):
        return self.pages

    def get_lexical_entries(self):
        return self.lexical_entries

    def get_lexical_senses(self):
        return self.lexical_senses

    def get_canonical_forms(self):
        return self.canonical_forms

    def pretty_print(self):
        for page_id in self.pages:
            print(self.pages[page_id])
            for entry_id in self.pages[page_id].get_entry_ids():
                print(self.lexical_entries[entry_id])
                for sense_id in self.lexical_entries[entry_id].get_sense_ids():
                    print(self.lexical_senses[sense_id])

    def add_page(self, page_id, page):
        #print("ADDING PAGE %s" % page_id)
        self.pages[page_id] = page

    def add_lexical_entry(self, entry_id, entry):
        #print("ADDING LE %s" % entry_id)
        self.lexical_entries[entry_id] = entry

    def add_lexical_sense(self, sense_id, sense):
        #print("ADDING SENSE %s" % sense_id)
        self.lexical_senses[sense_id] = sense

    def add_canonical_form(self, cf_id, gender):
        #print("ADDING CF %s" % cf_id)
        self.canonical_forms[cf_id] = gender

    def filter_wiki(self, trace=False):
        """
        Objectif : filtrage a posteriori des pages sans aucune LE de la bonne catégorie
                                     et  des LS correspondant à une LE de mauvaise catégorie
        """
        # à changer pour tracer
        if trace:
            print("BEFORE FILTERING")
            print("===================================")
            print(self.pages.keys())
            print(self.lexical_entries.keys())
            print(self.lexical_senses.keys())
            
            print("\n")
            for page_id in self.pages:
                print(page_id, self.pages[page_id].get_entry_ids())

            print("\n")

            for entry_id in self.lexical_entries:
                print((entry_id, self.lexical_entries[entry_id].get_sense_ids()))

            print("\n")

        #ajout des instances de LS au membre lexical_senses de chaque LE
        #et filtrage : on ne garde que les LS associés à des LE de la bonne catégorie
        new_lexical_senses = {}

        for entry_id in self.lexical_entries:
            le = self.lexical_entries[entry_id]
            sense_ids = le.get_sense_ids()
            new_sense_ids = []

            if sense_ids == None:
                continue
            else:
                for sense_id in sense_ids:
                    if sense_id not in self.lexical_senses:
                        continue
                    else:
                        sense = self.lexical_senses[sense_id]
                        new_lexical_senses[sense_id] = sense
                        new_sense_ids.append(sense_id)
            le.set_sense_ids(new_sense_ids)
        self.lexical_senses = new_lexical_senses


        #filtrage : on enlève les pages qui n'ont pas de LE correcte
        #           et les LE incorrectes pour chaque page
        new_pages = {}
        for page_id in self.pages:
            p = self.pages[page_id]
            # on ne garde que les LE non filtrées
            new_entry_ids = []
            for entry_id in p.get_entry_ids():
                if entry_id in self.lexical_entries:
                    new_entry_ids.append(entry_id)

            if len(new_entry_ids) > 0:
                p.entry_ids = new_entry_ids
                new_pages[page_id] = p
        self.pages = new_pages

        if trace:
            print("\nNB PAGES with at least one nominal entry %d" % len(self.pages.keys()))
            print("NB nominal Lexical Entries with %d" % len(self.lexical_entries.keys()))
            nb_mwe = len([ x for x in self.lexical_entries.items() if x[1].is_mwe == True])
            print("   among which %s are MWE" % nb_mwe)
            print("NB Senses of nominal lexical entries %d" % len(self.lexical_senses.keys()))
            print("\n")
            print("AFTER FILTERING")
            print("===================================")
            print(self.pages.keys())
            print(self.lexical_entries.keys())
            print(self.lexical_senses.keys())
            
            for page_id in self.pages:
                print(page_id, self.pages[page_id].get_entry_ids())
                print(f"{page_id} is multi entries page: {self.pages[page_id].is_multi_entries_page()}")

            print("\n")

            for entry_id in self.lexical_entries:
                print((entry_id, self.lexical_entries[entry_id].get_sense_ids()))
                print(f"{entry_id} is multi sense entry: {self.lexical_entries[entry_id].is_multi_senses_entry()}")
            print("\n")
            self.pretty_print()


class Parser:
    """
    A class used to represent an Animal

    ...

    Attributes
    ----------
    says_str : str
        a formatted string to print out what the animal says
    name : str
        the name of the animal
    sound : str
        the sound that the animal makes
    num_legs : int
        the number of legs the animal has (default 4)

    Methods
    -------
    says(sound=None)
        Prints the animals name and what sound it makes
    """
    def __init__(self, relations, trace, categories, rdf_types, seeds, lang,
                 file="fr_dbnary_ontolex.ttl", wiki=Wiktionary(), statistics_file="seeds_checked_V3.txt",
                 checked_seeds_file="seeds_checked_V3.txt", corpus_file="sequoia.deep_and_surf.parseme.frsemcor",
                 wiki_dump="wiki_dump.pkl", read=False, read_and_dump=False, check_seeds=False,
                 filter=False, get_seeds_from_corpus=False, statistics=False,
                 dump=None, mode='read', labels_to_ignore=[]):
        self.relations = relations
        self.dump = dump
        self.labels_to_ignore = labels_to_ignore
        self.mode = mode
        self.trace = trace
        self.categories = categories
        self.rdf_types = rdf_types
        self.wiki_dump = wiki_dump
        self.file = file
        self.wiki = wiki
        self.read_ = read
        self.checked_seeds_file = checked_seeds_file
        self.get_seeds_from_corpus_ = get_seeds_from_corpus
        self.statistics_ = statistics
        self.filter_ = filter
        self.check_seeds_ = check_seeds
        self.read_and_dump_ = read_and_dump
        self.seeds = seeds
        self.lang = lang
        self.statistics_file = statistics_file
        self.corpus_file = corpus_file

    def read(self):
        i = 0
        with open(self.file, "r", encoding="utf-8") as f:
            # while i < 1584351:
            #    f.readline()
            #    i += 1

            problematic_senses_ids = []

            # variables initialisation
            e1 = None
            relation = None
            e2 = None
            unallowed_cat = False
            unallowed_type = False
            unallowed_label = False
            last_line_of_paragraph = False
            rdf_type = None

            # dict temp data entry
            entry_temp_data = {}
            entry_temp_data["id"] = None
            entry_temp_data["is_mwe"] = False
            entry_temp_data["lemma"] = None
            entry_temp_data["pos"] = None
            entry_temp_data["synonym"] = set()
            entry_temp_data["cf_id"] = None
            entry_temp_data["sense_id"] = None

            # dict temp data sense
            sense_temp_data = {}
            sense_temp_data["id"] = None
            sense_temp_data["example"] = []
            sense_temp_data["definition"] = None
            sense_temp_data["synonym"] = set()
            sense_temp_data["labels"] = set()

            # read each line of the file
            while 1:
                line = f.readline()
                # premieres lignes definissant des @prefix
                if line.startswith('@'):
                    continue
                # exit from the while if EOF
                if not line:
                    break
                sline = line.strip()

                last_line_of_paragraph = False
                # if the line is not empty
                if sline:

                    # checks if this is the last line of a paragraph
                    if sline.endswith(" ."):  # or sline.endswith("@fr"):
                        last_line_of_paragraph = True

                    # if the line starts with a space or spaces then the e1 is the same as before
                    if line.startswith(" "):

                        # if the line ends with a dot then it is the last line of the paragraph
                        if last_line_of_paragraph:
                            delimiter = ' .'
                        else:
                            delimiter = ' ;'
                        # get the relation and e2
                        relation, e2 = line.strip().split(None, 1)
                        e2 = e2.strip()
                        if e2.startswith('[') and not last_line_of_paragraph:
                            while not e2.endswith('] ;') and not e2.endswith('] .'):
                                e2 += f.readline().strip()
                            if e2.endswith('] .'):
                                last_line_of_paragraph = True
                        e2 = e2.strip(delimiter).split(' , ')
                        e2 = [normalization_id(x, lang=self.lang) for x in e2]

                    # new paragraph
                    elif line.startswith(self.lang + ":") or line.startswith("<http://kaiko.getalp.org/dbnary/fra/"):
                        rdf_types = None
                        unallowed_type = False
                        unallowed_cat = False
                        unallowed_label = False
                        entry_temp_data["id"] = None
                        entry_temp_data["is_mwe"] = False
                        entry_temp_data["lemma"] = None
                        entry_temp_data["pos"] = None
                        entry_temp_data["synonym"] = set()
                        entry_temp_data["cf_id"] = None
                        entry_temp_data["sense_id"] = None
                        sense_temp_data["id"] = None
                        sense_temp_data["example"] = []
                        sense_temp_data["definition"] = None
                        sense_temp_data["synonym"] = set()
                        sense_temp_data["labels"] = set()

                        # if the is no ' ' in the line then there is only the e1 for the paragraph
                        if " " not in line:
                            # get the e1, relation and e2
                            e1 = line.strip()
                            e1 = normalization_id(e1, lang=self.lang)
                            relation = None
                            e2 = []
                        else:
                            # get the e1, relation and e2
                            e1, relation, e2 = line.strip().split(None, 2)
                            e2 = e2.strip()
                            if e2.startswith('[') and not last_line_of_paragraph:
                                while not e2.endswith('] ;') and not e2.endswith('] .'):
                                    e2 += f.readline().strip()
                                if e2.endswith('] .'):
                                    last_line_of_paragraph = True

                            e2 = e2.strip(' ;').strip(' .').split(' , ')
                            e2 = [normalization_id(x, lang=self.lang) for x in e2]
                            e1 = normalization_id(e1, lang=self.lang)
                    else:
                        print("WARNING: %s" % line)

                    # checks if there is a new rdf type for the paragraph
                    if relation == "rdf:type":
                        rdf_types = e2

                    # FILTERING
                    # if this line is an explicit triple
                    if rdf_types is not None:
                        # if the rdf type is note in those of interest then the paragraph is skipped
                        unallowed_type = True
                        for t in rdf_types:
                            if t in self.rdf_types:
                                unallowed_type = False
                                break

                    # checks if this is a line of interest or not
                    # NB: if last_line_of_paragraph, then we should go on even if not allowed_relations, to add the Page,  LE or LS
                    if relation == None or unallowed_type or (relation not in self.relations and not last_line_of_paragraph):
                        continue
                    # print the rdf triplet of the line
                    if self.trace:
                        print("e1: ", e1)
                        print("relation : ", relation)
                        print("e2: ", e2)
                        print("\n")

                    # print("e1: ", e1)
                    # print("relation : ", relation)
                    # print("e2: ", e2[0])
                    # print("\n")
                    # PROCESSING
                    # if rdf type is a page
                    if "dbnary:Page" in rdf_types:
                        page_id = e1
                        if relation == "dbnary:describes":
                            self.wiki.add_page(page_id=page_id, page=Page(id=page_id, entry_ids=e2))

                    # if rdf type is a lexical entry
                    elif "ontolex:LexicalEntry" in rdf_types:
                        entry_temp_data["id"] = e1
                        entry_temp_data["is_mwe"] = True if "ontolex:MultiWordExpression" in rdf_types else False

                        if relation == "rdfs:label":
                            entry_temp_data["lemma"] = e2[0].strip('"').removesuffix('"@fr')
                        if relation == "lexinfo:partOfSpeech":
                            entry_temp_data["pos"] = e2[0]
                            if e2[0] not in self.categories:
                                unallowed_cat = True
                        if relation == "dbnary:partOfSpeech":
                            unallowed_cat = True
                            for cat in self.categories:
                                if e2[0] == cat:
                                    # print("CAT: ", e2[0])
                                    unallowed_cat = False
                        if relation == "dbnary:synonym":
                            for synonym in e2:
                                entry_temp_data["synonym"].add(synonym)
                        if relation == "ontolex:canonicalForm":
                            entry_temp_data["cf_id"] = e2
                        if relation == "ontolex:sense":
                            entry_temp_data["sense_id"] = e2
                        # @@ on n'ajoute que si le filtrage categorie est ok
                        if last_line_of_paragraph and not unallowed_cat:
                            self.wiki.add_lexical_entry(entry_id=entry_temp_data["id"],
                                                        entry=lexicalEntry(id=entry_temp_data["id"],
                                                                           pos=entry_temp_data["pos"],
                                                                           lemma=entry_temp_data["lemma"],
                                                                           sense_ids=entry_temp_data["sense_id"],
                                                                           cf_ids=entry_temp_data["cf_id"],
                                                                           is_mwe=entry_temp_data["is_mwe"],
                                                                           synonyms=entry_temp_data["synonym"])
                                                        )
                            if self.trace:
                                for data in entry_temp_data:
                                    print(data, entry_temp_data[data])

                    # if rdf type is a lexical sense
                    elif "ontolex:LexicalSense" in rdf_types:
                        sense_temp_data["id"] = e1
                        # sense_temp_data["example"] = []
                        if relation == "skos:definition":
                            definition = extract_quote(text=e2[0])
                            # print(definition)
                            labels, definition = extract_labels(text=definition)
                            sense_temp_data["definition"] = definition
                            sense_temp_data["labels"] = labels
                            if labels is not None:
                                for l in labels:
                                    if l in self.labels_to_ignore:
                                        unallowed_label = True
                                        break


                            # TEST EXTRACTION LABELS AND DEFINITION
                            # print("\n")
                            # print(e2[0])
                            # print("DEFINITION: ", extract_quote(text=e2[0]))
                            # print("LABELS: ", labels)
                            # print("\n")
                        if relation == "skos:example":
                            # pass
                            sense_temp_data["example"].append(extract_quote(text=e2[0]))

                            # sense_temp_data["example"].append(e2[0])
                            # TEST EXTRACTION EXAMPLE
                            # print("\n")
                            # print("EXAMPLE: ", extract_quote(text=e2[0], left_delimiter='"', right_delimiter='@'))
                            # print("\n")
                        if relation == "dbnary:synonym":
                            for synonym in e2:
                                sense_temp_data["synonym"].add(synonym)
                        if last_line_of_paragraph and not unallowed_label:
                            sid = sense_temp_data["id"]
                            self.wiki.add_lexical_sense(sense_id=sid,
                                                        sense=lexicalSense(
                                                            id=sid, definition=sense_temp_data["definition"],
                                                            examples=sense_temp_data["example"],
                                                            synonyms=sense_temp_data["synonym"],
                                                            labels=sense_temp_data["labels"])
                                                        )
                            if self.trace:
                                for data in sense_temp_data:
                                    print(data, sense_temp_data[data])

                    # if rdf type is a canonical form
                    elif "ontolex:Form" in rdf_types:
                        if relation == "lexinfo:gender":
                            self.wiki.add_canonical_form(e1, e2[0])
        self.wiki.filter_wiki(trace=self.trace)
        # exit()
        print("READ TERMINATED")
        return self.wiki

    def filter(self):
        filtered_file = open("filtered_fr_dbnary_ontolex.ttl", "w", encoding="utf-8")
        with open(self.file, "r", encoding="utf-8") as f:

            # variables initialisation
            paragraph = []
            skip = False

            # read each line of the file
            while 1:
                line = f.readline()

                # premieres lignes definissant des @prefix
                if line.startswith('@'):
                    continue
                # exit from the while if EOF
                if not line:
                    break

                if line.startswith(self.lang + ":") or line.startswith("<"):
                    if not skip:
                        for l in paragraph:
                            filtered_file.write(l)
                    paragraph = []
                    skip = False
                if "rdf:type" in line:
                    # print("RDF_TYPE")
                    skip = True
                    for rdf_type in self.rdf_types:
                        if rdf_type in line:
                            skip = False
                if 'dbnary:partOfSpeech' in line or 'lexinfo:partOfSpeech' in line:
                    # print("POS")
                    skip = True
                    for cat in self.categories:
                        if cat in line:
                            # print("CATEGORY_"+cat)
                            skip = False

                paragraph.append(line)
        filtered_file.close()

    def read_and_dump(self):
        wiki = self.read()
        # Serialize the object to a file
        with open("wiki_dump.pkl", "wb") as file:
            pickle.dump(wiki, file)

    def check_seeds(self):

        cat = "nom"

        with open("wiki_dump.pkl", "rb") as file:
            self.wiki = pickle.load(file)

        if self.wiki.is_empty():
            self.wiki = self.read()

        # load the Excel file
        wb = openpyxl.load_workbook('seeds_ClassSemWikt_06_04-23.xlsx')

        # select the worksheet
        ws = wb.active

        # create new fill objects for each color
        # color for homograph nouns with at least one nominal entry having multiple senses
        violet_fill = PatternFill(start_color='8F00FF', end_color='8F00FF', fill_type='solid')
        # color for homograph nouns
        blue_fill = PatternFill(start_color='0000FF', end_color='0000FF', fill_type='solid')
        # color for nominal entries with more than one sense
        red_fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
        # color for unknown words
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        # color for nouns with only one sense and no homograph
        green_fill = PatternFill(start_color='00FF00', end_color='00FF00', fill_type='solid')

        pages = self.wiki.get_pages()
        entries = self.wiki.get_lexical_entries()
        senses = self.wiki.get_lexical_senses()
        cfs = self.wiki.get_canonical_forms()
        seen_words = set()

        g = open('ids_examples_missing.txt', 'w', encoding="utf-8")
        ids_examples_missing  = set()
        z = open('ids_definitions_missing.txt', 'w', encoding="utf-8")
        ids_definitions_missing = set()

        with open('seeds_to_be_manually_checked.txt', 'w', encoding="utf-8") as f:
            f.write("@SEEDS" + "\n")
            space_pattern = "\t"
            f.write("LEMMA" + space_pattern)
            f.write("ID_SENSE" + space_pattern)
            f.write("AMBIGUOUS" + space_pattern)
            f.write("SUPERSENSE" + space_pattern)
            f.write("ID_ENTRY_WIKI" + space_pattern)
            f.write("ID_SENSE_WIKI" + space_pattern)
            f.write("DEFINITION" + space_pattern)
            for i in range(1, 11):
                f.write(f"EXAMPLE_{i}" + space_pattern)
            f.write("\n")

            """
            f.write(str(page))
            for entry_id in pages[page_id].get_entry_ids():
                f.write(str(entries[entry_id]))
                for sense_id in entries[entry_id].get_sense_ids():
                    f.write(str(senses[sense_id]))
            """

            # loop through the cells in each row
            for col in ws.iter_cols():
                is_supersense = True
                for cell in col:
                    if is_supersense:
                        supersense = str(cell.value).strip().lower()
                        print("SUPERSENSE: ", supersense)
                        is_supersense = False
                    else:
                        # read the value from the cell
                        fill = green_fill
                        if cell.value is not None:
                            page_id = str(cell.value).strip()
                            if page_id in pages and page_id not in seen_words:
                                page = pages[page_id]
                                entry_ids = page.get_entry_ids()
                                multi_entries_page = False

                                # set the background color of the cell based on the corresponding data structure
                                if page.is_multi_entries_page():
                                    multi_entries_page = True
                                    fill = blue_fill
                                for entry_id in entry_ids:
                                    if entries[entry_id].is_multi_senses_entry() and multi_entries_page:
                                        fill = violet_fill
                                        break
                                    if entries[entry_id].is_multi_senses_entry() and not multi_entries_page:
                                        fill = red_fill
                                        break
                            else:
                                fill = yellow_fill

                            # WRITE THE TXT FILE FOR THE SEEDS WITH POLYSEMY OR HOMONYMY
                            if page_id in pages and page_id not in seen_words:
                                seen_words.add(page_id)
                                entry_num = 0
                                for entry_id in pages[page_id].get_entry_ids():
                                    if entry_id in entries:
                                        sense_num = 0
                                        entry_num += 1
                                        le = entries[entry_id]
                                        lemma = le.get_lemma()
                                        for sense_id in entries[entry_id].get_sense_ids():
                                            if sense_id in senses:
                                                sense_num += 1
                                                ls = senses[sense_id]
                                                new_sense_id = f"{lemma}__{cat}_{entry_num}_sens_{sense_num}"
                                                is_ambiguous = "yes" if not (fill == green_fill) else "no"
                                                f.write(f"{lemma}{space_pattern}")
                                                f.write(f"{new_sense_id}{space_pattern}")
                                                f.write(f"{is_ambiguous}{space_pattern}")
                                                f.write(f"{supersense}{space_pattern}")
                                                f.write(f"{entry_id}{space_pattern}")
                                                f.write(f"{sense_id}{space_pattern}")
                                                if ls.get_definition() is None:
                                                    f.write(f"DEF:{space_pattern}")
                                                else:
                                                    f.write(f"DEF:{ls.get_definition()}{space_pattern}")
                                                if ls.get_definition() is None:
                                                    ids_definitions_missing.add(sense_id)
                                                for example in ls.get_examples():
                                                    if example == "":
                                                        ids_examples_missing.add(sense_id)
                                                    else:
                                                        f.write(f"EX:{example}{space_pattern}")
                                                f.write("\n")
                            cell.fill = fill

        # save the modified Excel file
        wb.save('classified_seeds_ClassSemWikt_06_04-23.xlsx')
        for id in ids_examples_missing:
            g.write(id + "\n")
        g.close()
        for id in ids_definitions_missing:
            z.write(id + "\n")
        z.close()

    def statistics(self):
        seeds_statistical_analysis(self.statistics_file, SUPERSENSES, self.wiki_dump, revise=True)

    def get_seeds_from_corpus(self):
        new_seeds, lemma2supersenses = get_noun_entities_from_corpus(self.checked_seeds_file, self.corpus_file, self.wiki_dump)
        write_new_seeds(new_seeds, lemma2supersenses)

    def parse_file(self):
        if self.read_:
            print("READ")
            self.read()
        elif self.read_and_dump_:
            print("READ AND DUMP")
            self.read_and_dump()
        elif self.filter_:
            print("FILTER")
            self.filter()
        elif self.check_seeds_:
            print("CHECK_SEEDS")
            self.check_seeds()
        elif self.statistics_:
            print("STATISTICS")
            self.statistics()
        elif self.get_seeds_from_corpus_:
            print("GET_SEEDS")
            self.get_seeds_from_corpus()
        # return self.wiki

    def set_parser_mode(self, mode):
        if mode == "read":
            self.read_ = True
            self.read_and_dump_ = False
            self.filter_ = False
            self.check_seeds_ = False
            self.statistics_ = False
            get_seeds_from_corpus_ = False

        elif mode == "read_and_dump":
            self.read_ = False
            self.read_and_dump_ = True
            self.filter_ = False
            self.check_seeds_ = False
            self.statistics_ = False
            get_seeds_from_corpus_ = False

        elif mode == "filter":
            self.read_ = False
            self.read_and_dump_ = False
            self.filter_ = True
            self.check_seeds_ = False
            self.statistics_ = False
            get_seeds_from_corpus_ = False

        elif mode == "check_seeds":
            self.read_ = False
            self.read_and_dump_ = False
            self.filter_ = False
            self.check_seeds_ = True
            self.statistics_ = False
            get_seeds_from_corpus_ = False

        elif mode == "statistics":
            self.read_ = False
            self.read_and_dump_ = False
            self.filter_ = False
            self.check_seeds_ = False
            self.statistics_ = True
            get_seeds_from_corpus_ = False

        elif mode == "get_seeds_from_corpus":
            self.read_ = False
            self.read_and_dump_ = False
            self.filter_ = False
            self.check_seeds_ = False
            self.statistics_ = False
            get_seeds_from_corpus_ = True


# ------------------------------------------------------------

# supersenses acknowleged
SUPERSENSES = ['act', 'animal', 'artifact', 'attribute', 'body', 'cognition',
               'communication', 'event', 'feeling', 'food', 'institution', 'act*cognition',
               'object', 'possession', 'person', 'phenomenon', 'plant', 'artifact*cognition',
               'quantity', 'relation', 'state', 'substance', 'time', 'groupxperson']


def seeds_statistical_analysis(file, supersenses, wiki_file, revise=False):
    seeds = open(file, 'r', encoding="utf-8")
    seeds_to_be_revised = open("seeds_to_be_revised.txt", 'w', encoding="utf-8")
    headers = seeds.readline().split('\t')
    headers = headers[:-1] if headers[-1] == "\n" else headers
    headers = [header.strip("\n") for header in headers ]
    # print(headers)
    with open(wiki_file, "rb") as file:
        wiki = pickle.load(file)
    domain_labels = []
    with open("domain_labels.txt", "r", encoding="utf-8") as domain_file:
        while 1:
            line = domain_file.readline()
            if not line:
                break
            domain_labels.append(line.strip().strip("\n"))
    # print(domain_labels)
    # print(len(domain_labels))

    nb_senses_with_labels = 0
    nb_senses_with_synonyms = 0
    lemmas = set()
    non_ambiguous_lemmas = set()
    ambiguous_lemmas = set()
    domain_ambiguity = set()
    homonymy = set()
    polysemy = set()
    homonymy_and_polysemy = set()
    has_homonymy = False
    has_polysemy = False
    has_homonymy_and_polysemy = False
    has_domain_ambiguity = False
    senses_nb = 0
    non_ambiguous_senses_nb = 0
    ambiguous_senses_nb = 0
    supersense2non_ambiguous_lemma = defaultdict(int)
    supersense2non_ambiguous_sense = defaultdict(int)
    supersense2ambiguous_lemma = defaultdict(int)
    supersense2ambiguous_sense = defaultdict(int)
    lemma = ""
    senses = []
    supersenses_for_lemma = set()
    labels_for_lemma = set()
    nb_supersenses_for_lemma = 0
    nb_supersenses_for_ambiguous_lemma = 0
    nb_supersenses_for_non_ambiguous_lemma = 0
    first_line = True
    last_line = False
    while 1:
        line = seeds.readline()
        if not line:
            last_line = True
        else:
            sense = line.strip().split('\t')
            sense = sense[:-1] if sense[-1] == "\n" else sense
            lemmas.add(sense[0])
            senses_nb += 1
        if sense[0] == lemma or first_line:
            if sense[4]:
                labels_for_sense = wiki.lexical_senses[sense[4]].labels
            else:
                labels_for_sense = set()

            if labels_for_sense:
                labels_for_lemma = labels_for_lemma.union(labels_for_sense)
            supersenses_for_lemma.add(sense[2])
            senses.append(sense)
            if first_line:
                lemma = sense[0]
                first_line = False
        if sense[0] != lemma or last_line:
            # process every line for previous lemma
            is_ambiguous = False
            is_to_be_revised = False
            nb_supersenses_for_lemma += len(supersenses_for_lemma)
            for label in labels_for_lemma:
                if any(label == domain for domain in domain_labels):
                    has_domain_ambiguity = True
                    # print(next(domain for domain in domain_labels if label == domain ))
                    break

            for line in senses:
                if line[1] == "yes":
                    is_ambiguous = True
                if wiki.lexical_senses[line[4]].labels:
                    if len(wiki.lexical_senses[line[4]].labels) >= 1:
                        nb_senses_with_labels += 1
                if wiki.lexical_senses[line[4]].synonyms:
                    if len(wiki.lexical_senses[line[4]].synonyms) >= 1:
                        nb_senses_with_synonyms += 1

            if is_ambiguous:
                nb_supersenses_for_ambiguous_lemma += len(supersenses_for_lemma)
                ambiguous_lemmas.add(lemma)
                for ss in list(supersenses_for_lemma):
                    supersense2ambiguous_lemma[ss] += 1
                has_homonymy = wiki.pages[lemma].is_multi_entries_page()
                for line in senses:
                    ambiguous_senses_nb += 1
                    supersense2ambiguous_sense[line[2]] += 1
                    if has_polysemy:
                        continue
                    else:
                        has_polysemy = wiki.lexical_entries[line[3]].is_multi_senses_entry()
                has_homonymy_and_polysemy = has_polysemy and has_homonymy
                if has_polysemy:
                    polysemy.add(lemma)
                if has_homonymy:
                    homonymy.add(lemma)
                if has_homonymy_and_polysemy:
                    homonymy_and_polysemy.add(lemma)
                if has_domain_ambiguity:
                    domain_ambiguity.add(lemma)


            else:
                if senses:
                    nb_supersenses_for_non_ambiguous_lemma += len(supersenses_for_lemma)
                    non_ambiguous_lemmas.add(lemma)
                    for ss in list(supersenses_for_lemma):
                        supersense2non_ambiguous_lemma[ss] += 1

                    for line in senses:
                        non_ambiguous_senses_nb += 1
                        supersense2non_ambiguous_sense[line[2]] += 1

            senses = [sense]
            lemma = sense[0]
            supersenses_for_lemma = set()
            labels_for_lemma = set()
            # print(sense)
            supersenses_for_lemma.add(sense[2])
            has_homonymy = False
            has_polysemy = False
            has_homonymy_and_polysemy = False
            has_domain_ambiguity = False

        if last_line:
            break
    seeds.close()
    seeds_to_be_revised.close()

    # print statistics for the seeds
    print(f"nb_lemmes_wiki = {len(wiki.pages)}")
    print(f"nb_sens_wiki = {len(wiki.lexical_senses)}")
    print(f"number of lemmas = {len(lemmas)}")
    print(f"number of senses = {senses_nb}")
    print(f"number of ambiguous lemmas = {len(ambiguous_lemmas)}")
    print(f"number of non ambiguous lemmas = {len(non_ambiguous_lemmas)}")
    print(f"proportion of ambiguous lemmas = {len(ambiguous_lemmas)/len(lemmas)}")
    print(f"number of senses with labels = {nb_senses_with_labels}")
    print(f"number of senses with synonyms = {nb_senses_with_synonyms}")

    for supersense in supersenses:
        print(
            f"number of senses for supersense {supersense} = {supersense2non_ambiguous_sense[supersense] + supersense2ambiguous_sense[supersense]}")


    print(f"nb moyen de lemmes non ambigus par supersense = {sum(supersense2non_ambiguous_lemma.values())/len(supersenses)}")
    print(
        f"nb moyen de lemmes ambigus par supersense = {sum(supersense2ambiguous_lemma.values()) / len(supersenses)}")

    print(f"nb moyen de supersenses par lemme = {nb_supersenses_for_lemma/len(lemmas)}")
    print(f"nb moyen de supersenses par lemme ambigu = {nb_supersenses_for_ambiguous_lemma/len(ambiguous_lemmas)}")
    print(f"nb moyen de supersenses par lemme non ambigu = {nb_supersenses_for_non_ambiguous_lemma/len(non_ambiguous_lemmas)}")

    print(f"nb lemmes ambigus avec homonimie = {len(homonymy)}")
    print(f"nb lemmes ambigus avec polysémie = {len(polysemy)}")
    print(f"nb lemmes ambigus avec homonimie et polysémie = {len(homonymy_and_polysemy)}")

    print(f"nb lemmes ambigus avec une étiquette de domaine = {len(domain_ambiguity)}")
    print(f"proportion de lemmes ambigus avec une étiquette de domaine = {len(domain_ambiguity)/len(lemmas)}")
    print(f"proportion de lemmes ambigus avec une étiquette de domaine parmi les lemmes ambigus = {len(domain_ambiguity) / len(ambiguous_lemmas)}")


def get_supersenses_from_sequoia_lemma_in_sentence(lemma, sentence):
    supersenses_to_be_returned = set()
    index2supersense = {}
    index_patterns = [f"{i}:" for i in range(1,20)]
    indices = [f"{i}" for i in range(1, 20)]

    for word in sentence:
        if word[2] == lemma:
            if ';' in word[11]:
                supersenses = word[11].strip().split(";")
            else:
                supersenses = [word[11]]

            for supersense in supersenses:

                if any(pattern in supersense for pattern in index_patterns):
                    num = next(pattern for pattern in index_patterns if pattern in supersense)
                    supersenses_to_be_returned.add(supersense.strip(num))
                    index2supersense[num.strip(':')] = supersense.strip(num)

                elif any(index in supersense for index in indices):
                    num = next(index for index in indices if index in supersense)
                    if num in index2supersense:
                        supersenses_to_be_returned.add(index2supersense[num])
                        index2supersense[num.strip(':')] = supersense.strip(num)
                    else:
                        for word in sentence:
                            if ';' in word[11]:
                                ss = word[11].strip().split(";")
                            else:
                                ss = [word[11]]
                            for s in ss:
                                if any(pattern in s for pattern in index_patterns):
                                    num = next(pattern for pattern in index_patterns if pattern in s)
                                    supersenses_to_be_returned.add(s.strip(num))
                                    index2supersense[num.strip(':')] = s.strip(num)
                else:
                    supersenses_to_be_returned.add(supersense)
    # print(list(supersenses_to_be_returned))
    return list(supersenses_to_be_returned)


def get_noun_entities_from_corpus(old_seeds_file, sequoia_file, wiki_file):
    with open(wiki_file, "rb") as file:
        wiki = pickle.load(file)
    seen_words = set()
    pages = wiki.pages
    lexical_entries = wiki.lexical_entries
    lexical_senses = wiki.lexical_senses
    ids_definitions_missing = set()
    ids_examples_missing = set()
    lemma2supersenses = defaultdict(set)

    with open(old_seeds_file, "r", encoding="utf-8") as z:
        z.readline()
        while 1:
            line = z.readline()
            if not line:
                break
            sline = line.strip().strip("\n").split("\t")
            seen_words.add(sline[0])

    corpus_name = "sequoia"
    new_seeds = open(f"seeds_from_{corpus_name}.txt", 'w', encoding="utf-8")
    corpus_name_file = f"seeds_from_{corpus_name}.txt"
    corpus = open(sequoia_file, 'r', encoding="utf-8")
    space_pattern = "\t"

    new_seeds.write("LEMMA" + space_pattern)
    new_seeds.write("DEFINITION" + space_pattern)
    for i in range(1, 11):
        new_seeds.write(f"EXAMPLE_{i}" + space_pattern)
    new_seeds.write("\n")
    lemma2definitions = defaultdict(list)
    lemma2examples = defaultdict(list)
    lemmas = set()
    last_line = False
    very_first_word = True
    words = []
    while 1:
        line = corpus.readline()
        if line.startswith("#"):
            continue
        if not line:
            last_line = True

        if line.strip():
            sline = line.strip().strip("\n").split("\t")
            if (sline[0] == "1" or last_line) and not very_first_word:
                for word in words:
                    if word[3] == "N":
                        lemma = word[2]
                        lemmas.add(lemma)
                        supersenses = get_supersenses_from_sequoia_lemma_in_sentence(lemma, words)
                        # print(supersenses)
                        for supersense in supersenses:
                            if supersense != "*":
                                lemma2supersenses[lemma].add(supersense)
                # get current sentence 1st word
                words = [sline]
            else:
                sline = line.strip().strip("\n").split("\t")
                words.append(sline)
                very_first_word = False
        if last_line:
            break

    for lemma in list(lemmas):
        page_id = lemma
        if page_id in pages and page_id not in seen_words:
            seen_words.add(page_id)
            entry_num = 0
            for entry_id in pages[page_id].get_entry_ids():
                if entry_id in lexical_entries:
                    sense_num = 0
                    entry_num += 1
                    le = lexical_entries[entry_id]
                    lemma = le.get_lemma()
                    for sense_id in lexical_entries[entry_id].get_sense_ids():
                        if sense_id in lexical_senses:
                            sense_num += 1
                            ls = lexical_senses[sense_id]
                            new_sense_id = f"{lemma}__nom_{entry_num}_sens_{sense_num}"
                            is_ambiguous = "yes" if (pages[page_id].is_multi_entries_page() or lexical_entries[entry_id].is_multi_senses_entry()) else "no"
                            new_seeds.write(f"{lemma}{space_pattern}")

                            if ls.get_definition() is None:
                                ids_definitions_missing.add(sense_id)
                                # new_seeds.write(f"DEF:{temp_definitions[sense_id]}{space_pattern}")
                                new_seeds.write(f"DEF:{''}{space_pattern}")
                            else:
                                new_seeds.write(f"DEF:{ls.get_definition()}{space_pattern}")

                            for example in ls.get_examples():
                                if example == "":
                                    ids_examples_missing.add(sense_id)
                                else:
                                    new_seeds.write(f"EX:{example}{space_pattern}")
                            new_seeds.write("\n")

    new_seeds.close()
    corpus.close()
    print("DEF MANQUANTES: ", len(ids_definitions_missing))
    print("EX MANQUANTS: ", len(ids_examples_missing))
    print(f"len(lemmas) = {len(lemmas)}")
    return corpus_name_file, lemma2supersenses


def write_new_seeds(corpus_seeds_file, lemma2supersenses):
    new_seeds = open("new_seeds.txt", 'w', encoding="utf-8")
    corpus = open(corpus_seeds_file, 'r', encoding="utf-8")
    space_pattern = "\t"
    new_seeds.write("LEMMA" + space_pattern)
    new_seeds.write("SUPERSENSES" + space_pattern)
    new_seeds.write("DEFINITION" + space_pattern)
    for i in range(1, 11):
        new_seeds.write(f"EXAMPLE_{i}" + space_pattern)
    new_seeds.write("\n")
    nb_examples = 10
    senses = []
    lemma = ""
    first_line = True
    last_line = False
    while 1:
        line = corpus.readline()

        if line.startswith("LEMMA"):
            continue
        if not line:
            last_line = True

        sense = line.strip().split('\t')
        sense = sense[:-1] if sense[-1] == "\n" else sense
        if (sense[0] == lemma or first_line) and not last_line:
            senses.append(sense)
            if first_line:
                lemma = sense[0]
                first_line = False
        if sense[0] != lemma or last_line:
            # process every line for previous lemma
            lemma_first_time = True
            supersenses_not_written = True
            for line in senses:
                lemma = line[0]
                supersenses = list(lemma2supersenses[lemma])
                examples = []
                definition = line[1]
                for ex in line[2:]:
                    examples.append(ex)

                new_seeds.write(f"{lemma}{space_pattern}") if lemma_first_time else new_seeds.write(f"{space_pattern}")

                if supersenses_not_written:
                    supersenses_list = ""
                    for ss in supersenses:
                       supersenses_list += f"{ss};"
                    supersenses_list = supersenses_list.strip(';')
                    new_seeds.write(f"{supersenses_list}{space_pattern}")
                else:
                    new_seeds.write(f"{space_pattern}")
                supersenses_not_written = False
                new_seeds.write(f"{definition}{space_pattern}")
                for ex in examples:
                    new_seeds.write(f"{ex}{space_pattern}")
                new_seeds.write("\n")
                lemma_first_time = False

            senses = [sense]
            lemma = sense[0]

        if last_line:
            break

    new_seeds.close()
    corpus.close()


def from_table_to_text_file(table_file, supersenses_to_consider):
    table_name = ""
    for c in table_file:
        if c == ".":
            break
        else:
            table_name += c

    with open(f'{table_name}.txt', 'w', encoding="utf-8") as file:
        wb = openpyxl.load_workbook(table_file)
        ws = wb.active
        headers = []

        first_line = ""
        for cell in ws.iter_cols(min_row=1, max_row=1, values_only=True):
            if cell[0] is not None:
                headers.append(str(cell[0]).strip().lower())
                first_line += f"{str(cell[0]).strip().lower()}\t"
        first_line = first_line.strip("\t")
        file.write(first_line+"\n")


        for row in ws.iter_rows(min_row=2):
            line = ""
            row_supersense = None
            for j, cell in enumerate(row):
                if j < len(headers):
                    if cell.value is None:
                        line += f"\t"
                    else:
                        value = str(cell.value).strip()
                        if headers[j].lower() == 'supersense':
                            if value in supersenses_to_consider:
                                row_supersense = value
                                line += f"{value}\t"
                            else:
                                break
                        else:
                            line += f"{value}\t"
            if row_supersense:
                line = line.strip("\t")
                file.write(f"{line}\n")
        print(headers)


def get_sense_id_and_entry_id_from_lemma_and_defition(lemma, definition, wiki):
    for entry_id in wiki.pages[lemma].entry_ids:
        for sense_id in wiki.lexical_entries[entry_id].sense_ids:
            sense = wiki.lexical_senses[sense_id]
            if sense.definition == definition.replace("DEF:", ""):
                le_id = entry_id
                ls_id = sense_id
    return ls_id, le_id


def add_sense_id_and_entry_id_to_seeds_checked_V2(file="seeds_checked_V2.txt", wiki_file="wiki_dump.pkl"):
    old_file = open(file, 'r', encoding="utf-8")
    new_file = open("seeds_checked_V3.txt", 'w', encoding="utf-8")

    with open(wiki_file, "rb") as file:
        wiki = pickle.load(file)

    new_file.write(old_file.readline().replace("\tquestion", "").replace("id_sense\t", ""))
    while 1:
        line = old_file.readline()
        if not line:
            break
        sline = line.strip().strip("\n").split("\t")
        print(sline)
        lemma = sline[0]
        if len(sline) <= 7:
            definition = "DEF:"
        else:
            definition = sline[7]
        if not sline[2] == "oui":
            if sline[1] == "" and sline[5] == "" and sline[6] == "":
                sense_id, entry_id = get_sense_id_and_entry_id_from_lemma_and_defition(lemma, definition, wiki)
                new_sline = [el for el in sline]
                new_sline[5] = entry_id
                new_sline[6] = sense_id
                new_sline.pop(1)
                new_line = ""
                for el in new_sline:
                    new_line += f"{el}\t"
                new_line = new_line.strip("\t")
                new_line += "\n"
                new_line = new_line.replace("\t\t\t", "\t\t")
                new_file.write(new_line)
            else:
                new_sline = sline
                new_sline.pop(1)
                new_line = ""
                for el in new_sline:
                    new_line += f"{el}\t"
                new_line = new_line.strip("\t")
                new_line += "\n"
                new_line = new_line.replace("\t\t", "\t")
                new_file.write(new_line)

    new_file.close()
    old_file.close()


def get_natural_selection_of_senses_from_corpus(wiki_file="wiki_dump.pkl", seeds_file="seeds_checked_V3.txt"):
    with open(wiki_file, "rb") as file:
        wiki = pickle.load(file)
    seeds = open(seeds_file, 'r', encoding="utf-8")
    senses = open("baseline_random_senses_to_annotate_from_wiktionary.txt", 'w', encoding="utf-8")
    seeds_senses = {}

    first_line = seeds.readline()
    headers = first_line.strip().strip("\n").split("\t")
    senses.write(first_line.replace("\tambiguous", "").replace("\tdefinition\t", "\tlabels\tsynonyms\tdefinition\t"))

    while 1:
        line = seeds.readline()
        if not line:
            break
        if line.strip():
            sline = line.strip().strip("\n").split("\t")
            for i, header in enumerate(headers):
                if header == "supersense":
                    supersense = sline[i]
                if header == "id_sense_wiki":
                    sense_id = sline[i]
            seeds_senses[sense_id] = supersense

    lexical_senses_ids = list(wiki.lexical_senses.keys())
    nb_iterations = 50000
    sample_size = 1000
    for i in range(nb_iterations):
        random.shuffle(lexical_senses_ids)
    random_senses_ids = random.sample(lexical_senses_ids, sample_size)

    for lemma in wiki.pages:
        for entry_id in wiki.pages[lemma].entry_ids:
            if wiki.lexical_entries[entry_id].sense_ids:
                for sense_id in wiki.lexical_entries[entry_id].sense_ids:
                    if sense_id in random_senses_ids:

                        labels = ""
                        if wiki.lexical_senses[sense_id].labels:
                            for label in list(wiki.lexical_senses[sense_id].labels):
                                labels += (label + ";")
                        labels = labels.strip(";")

                        synonyms = ""
                        if wiki.lexical_senses[sense_id].synonyms:
                            for synonym in list(wiki.lexical_senses[sense_id].synonyms):
                                synonyms += (synonym + ";")
                        synonyms = synonyms.strip(";")

                        definition = wiki.lexical_senses[sense_id].definition

                        examples = ""
                        for example in wiki.lexical_senses[sense_id].examples:
                            examples += (example + "\t")
                        examples = examples.strip("\t")

                        if sense_id in seeds_senses:
                            senses.write(
                                f"{lemma}\t{seeds_senses[sense_id]}\t{entry_id}\t{sense_id}\t{labels}\t{synonyms}\t{definition}\t{examples}\n")
                        else:
                            senses.write(
                                f"{lemma}\t{'*'}\t{entry_id}\t{sense_id}\t{labels}\t{synonyms}\t{definition}\t{examples}\n")

    seeds.close()
    senses.close()


get_natural_selection_of_senses_from_corpus()
